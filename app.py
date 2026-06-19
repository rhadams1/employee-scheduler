"""
Ice Line Employee Scheduler - Flask Backend
Production-ready version with proper error handling and configuration
"""

import os
import calendar as cal
import logging
from datetime import datetime, timedelta, date
from io import BytesIO
from functools import wraps

from dotenv import load_dotenv
load_dotenv()

from flask import Flask, jsonify, request, send_file, render_template, g, session, redirect
import sqlite3

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from auth import (
    manager_required, employee_required,
    login_user, logout_user, verify_login,
)

# =============================================================================
# CONFIGURATION
# =============================================================================

class Config:
    """Application configuration"""
    SECRET_KEY = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
    DATABASE = os.environ.get('DATABASE_PATH', 'schedule.db')
    DEBUG = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    BACKUP_IMPORT_ENABLED = os.environ.get('BACKUP_IMPORT_ENABLED', 'false').lower() == 'true'
    PDF_RENDER_BASE_URL = os.environ.get('PDF_RENDER_BASE_URL', 'http://127.0.0.1:5001')
    PDF_RENDER_TIMEOUT_MS = int(os.environ.get('PDF_RENDER_TIMEOUT_MS', '10000'))
    AUTH_ENABLED = os.environ.get('AUTH_ENABLED', 'false').lower() == 'true'
    # Session cookie hardening. HTTPS comes from the Cloudflare tunnel in prod;
    # SESSION_COOKIE_SECURE stays off for LAN/dev unless explicitly enabled.
    SESSION_COOKIE_HTTPONLY = True
    SESSION_COOKIE_SAMESITE = 'Lax'
    SESSION_COOKIE_SECURE = os.environ.get('SESSION_COOKIE_SECURE', 'false').lower() == 'true'
    PERMANENT_SESSION_LIFETIME = timedelta(days=30)  # employee "remember this device"
    AUTH_MAX_FAILED = int(os.environ.get('AUTH_MAX_FAILED', '5'))
    AUTH_LOCKOUT_MINUTES = int(os.environ.get('AUTH_LOCKOUT_MINUTES', '15'))

    # Schedule configuration
    WEEK_START_DAY = 'wednesday'  # Schedule week starts on Wednesday
    DEFAULT_OFFICE_OPEN = '8:00 AM'
    DEFAULT_OFFICE_CLOSE = '10:00 PM'
    
    # Default employees (only used on first run)
    DEFAULT_EMPLOYEES = [
        ('Bob Adams', '610-505-6322', 'manager', 1),
        ('Dave Hendricks', '484-459-8620', 'manager', 2),
        ('Zak Reilly', '267-247-2955', 'zak', 1),
        ('Ava Hawthorne', '(267) 738-4698', 'staff', 1),
        ('Marisa Fullerton', '(215) 252-6544', 'staff', 2),
        ('Nate Bailey', '(609) 832-9499', 'staff', 3),
        ('Lilli Binns', '', 'staff', 4),
        ('Olivia Binns', '', 'staff', 5),
        ('Hunter Haas', '(484) 631-5469', 'staff', 6),
        ('Lena Sturz', '', 'staff', 7),
    ]


# =============================================================================
# APPLICATION FACTORY
# =============================================================================

def create_app(config_class=Config):
    """Application factory pattern"""
    app = Flask(__name__, 
                static_folder='static', 
                template_folder='templates')
    
    app.config.from_object(config_class)
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if app.config['DEBUG'] else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Register blueprints/routes
    register_routes(app)

    # Register CLI commands (bootstrap credentials)
    from cli import register_cli
    register_cli(app)

    # Register error handlers
    register_error_handlers(app)
    
    # Setup database
    with app.app_context():
        init_db()
    
    return app


# =============================================================================
# DATABASE
# =============================================================================

def get_db():
    """Get database connection for current request"""
    if 'db' not in g:
        g.db = sqlite3.connect(Config.DATABASE, timeout=5.0)
        g.db.row_factory = sqlite3.Row
        # WAL mode + busy_timeout: needed because we run 9 gunicorn workers,
        # auto-save fires every 1.5s, and saves do delete-then-insert per week
        g.db.execute('PRAGMA journal_mode = WAL')
        g.db.execute('PRAGMA busy_timeout = 5000')
        g.db.execute('PRAGMA foreign_keys = ON')
    return g.db


def close_db(e=None):
    """Close database connection"""
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    """Initialize database schema and default data"""
    conn = sqlite3.connect(Config.DATABASE)
    cursor = conn.cursor()
    
    # Create tables
    cursor.executescript('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT DEFAULT '',
            section TEXT NOT NULL CHECK(section IN ('manager', 'zak', 'staff')),
            sort_order INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        
        CREATE TABLE IF NOT EXISTS schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            week_start DATE NOT NULL UNIQUE,
            week_title TEXT NOT NULL,
            version INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        
        CREATE TABLE IF NOT EXISTS shifts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER NOT NULL,
            employee_id INTEGER NOT NULL,
            day_index INTEGER NOT NULL CHECK(day_index >= 0 AND day_index <= 6),
            time_in TEXT,
            time_out TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (schedule_id) REFERENCES schedules(id) ON DELETE CASCADE,
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
            UNIQUE(schedule_id, employee_id, day_index)
        );
        
        CREATE TABLE IF NOT EXISTS office_hours (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER NOT NULL,
            day_index INTEGER NOT NULL CHECK(day_index >= 0 AND day_index <= 6),
            time_in TEXT,
            time_out TEXT,
            FOREIGN KEY (schedule_id) REFERENCES schedules(id) ON DELETE CASCADE,
            UNIQUE(schedule_id, day_index)
        );
        
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER NOT NULL,
            day_index INTEGER NOT NULL CHECK(day_index >= 0 AND day_index <= 6),
            event_text TEXT NOT NULL,
            FOREIGN KEY (schedule_id) REFERENCES schedules(id) ON DELETE CASCADE
        );
        
        CREATE TABLE IF NOT EXISTS employee_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            week_start DATE NOT NULL,
            note TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
            UNIQUE(employee_id, week_start)
        );
        
        -- Indexes for performance
        CREATE INDEX IF NOT EXISTS idx_shifts_schedule ON shifts(schedule_id);
        CREATE INDEX IF NOT EXISTS idx_shifts_employee ON shifts(employee_id);
        CREATE INDEX IF NOT EXISTS idx_schedules_week ON schedules(week_start);
        CREATE INDEX IF NOT EXISTS idx_employees_active ON employees(active, section);
    ''')

    # Migrations: add columns to existing DBs that pre-date them
    cursor.execute("PRAGMA table_info(schedules)")
    schedule_cols = {row[1] for row in cursor.fetchall()}
    if 'version' not in schedule_cols:
        cursor.execute("ALTER TABLE schedules ADD COLUMN version INTEGER NOT NULL DEFAULT 1")
        logging.info("Migrated: added schedules.version column")

    # Auth columns (Phase 1 auth foundation) — additive, safe to re-run on existing DBs
    cursor.execute("PRAGMA table_info(employees)")
    employee_cols = {row[1] for row in cursor.fetchall()}
    if 'role' not in employee_cols:
        cursor.execute("ALTER TABLE employees ADD COLUMN role TEXT NOT NULL DEFAULT 'employee'")
        logging.info("Migrated: added employees.role column")
    if 'username' not in employee_cols:
        cursor.execute("ALTER TABLE employees ADD COLUMN username TEXT")
    if 'password_hash' not in employee_cols:
        cursor.execute("ALTER TABLE employees ADD COLUMN password_hash TEXT")
    if 'pin_hash' not in employee_cols:
        cursor.execute("ALTER TABLE employees ADD COLUMN pin_hash TEXT")
    if 'failed_attempts' not in employee_cols:
        cursor.execute("ALTER TABLE employees ADD COLUMN failed_attempts INTEGER NOT NULL DEFAULT 0")
    if 'locked_until' not in employee_cols:
        cursor.execute("ALTER TABLE employees ADD COLUMN locked_until TIMESTAMP")
    cursor.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_employees_username "
        "ON employees(username) WHERE username IS NOT NULL"
    )

    # Seed default employees if table is empty
    cursor.execute('SELECT COUNT(*) FROM employees')
    if cursor.fetchone()[0] == 0:
        cursor.executemany(
            'INSERT INTO employees (name, phone, section, sort_order) VALUES (?, ?, ?, ?)',
            Config.DEFAULT_EMPLOYEES
        )
        logging.info("Initialized default employees")

    conn.commit()
    conn.close()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_week_dates(week_start_str):
    """Generate day info for the week starting from Wednesday"""
    week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
    day_names = ['Wed', 'Thurs', 'Fri', 'Sat', 'Sun', 'Mon', 'Tues']
    wed = week_start + timedelta(days=2)  # Monday + 2 = Wednesday
    
    return [
        {
            'name': name,
            'date': f"{(wed + timedelta(days=i)).month}/{(wed + timedelta(days=i)).day}",
            'fullDate': (wed + timedelta(days=i)).isoformat(),
            'isWeekend': name in ['Sat', 'Sun']
        }
        for i, name in enumerate(day_names)
    ]


def parse_time_to_minutes(time_str):
    """Convert time string like '9:00 AM' to minutes since midnight"""
    if not time_str or time_str in ('-', 'CLOSE', 'CLOSED'):
        return None
    try:
        time_str = time_str.strip().upper()
        if 'AM' in time_str or 'PM' in time_str:
            parts = time_str.replace('AM', '').replace('PM', '').strip().split(':')
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 else 0
            if 'PM' in time_str and hour != 12:
                hour += 12
            if 'AM' in time_str and hour == 12:
                hour = 0
            return hour * 60 + minute
    except (ValueError, IndexError):
        return None
    return None


def format_week_title(wed_date):
    """Format week title like 'December 31st, 2025'"""
    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    
    day = wed_date.day
    if day in [1, 21, 31]:
        suffix = 'st'
    elif day in [2, 22]:
        suffix = 'nd'
    elif day in [3, 23]:
        suffix = 'rd'
    else:
        suffix = 'th'
    
    return f"{months[wed_date.month - 1]} {day}{suffix}, {wed_date.year}"


def get_current_week_start():
    """Calculate the Monday of the current schedule week"""
    today = date.today()
    weekday = today.weekday()  # 0=Mon, 1=Tue, 2=Wed, etc.
    
    # Schedule runs Wed-Tues
    # If Mon(0) or Tue(1), we're still in the previous week's schedule
    if weekday < 2:
        days_back = 7 + weekday
    else:
        days_back = weekday
    
    monday = today - timedelta(days=days_back)
    return monday.isoformat()


def get_or_create_schedule(week_start_str):
    """Get existing schedule or create new one"""
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute('SELECT * FROM schedules WHERE week_start = ?', (week_start_str,))
    schedule = cursor.fetchone()
    
    if not schedule:
        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        wed = week_start + timedelta(days=2)
        week_title = format_week_title(wed)
        
        cursor.execute(
            'INSERT INTO schedules (week_start, week_title) VALUES (?, ?)',
            (week_start_str, week_title)
        )
        schedule_id = cursor.lastrowid
        
        # Create default office hours
        for i in range(7):
            cursor.execute(
                'INSERT INTO office_hours (schedule_id, day_index, time_in, time_out) VALUES (?, ?, ?, ?)',
                (schedule_id, i, Config.DEFAULT_OFFICE_OPEN, Config.DEFAULT_OFFICE_CLOSE)
            )
        
        db.commit()
        cursor.execute('SELECT * FROM schedules WHERE id = ?', (schedule_id,))
        schedule = cursor.fetchone()
    
    return dict(schedule)


def build_schedule_response(week_start_str):
    """Build complete schedule data for API response"""
    db = get_db()
    cursor = db.cursor()
    
    schedule = get_or_create_schedule(week_start_str)
    schedule_id = schedule['id']

    # Show employees who are currently active OR who have shifts on this week
    # (so past weeks still show seasonal staff who worked then, even after they're hidden)
    cursor.execute('''
        SELECT * FROM employees
        WHERE active = 1
           OR id IN (SELECT DISTINCT employee_id FROM shifts WHERE schedule_id = ?)
        ORDER BY
            CASE section WHEN 'manager' THEN 1 WHEN 'zak' THEN 2 WHEN 'staff' THEN 3 END,
            sort_order
    ''', (schedule_id,))
    employees = [dict(row) for row in cursor.fetchall()]
    
    # Get shifts as lookup dictionary
    cursor.execute('SELECT * FROM shifts WHERE schedule_id = ?', (schedule_id,))
    shifts = {
        (row['employee_id'], row['day_index']): {'in': row['time_in'], 'out': row['time_out']}
        for row in cursor.fetchall()
    }
    
    # Get office hours
    cursor.execute('SELECT * FROM office_hours WHERE schedule_id = ? ORDER BY day_index', (schedule_id,))
    oh_dict = {row['day_index']: row for row in cursor.fetchall()}
    office_hours = [
        {'in': oh_dict[i]['time_in'], 'out': oh_dict[i]['time_out']} if i in oh_dict
        else {'in': Config.DEFAULT_OFFICE_OPEN, 'out': Config.DEFAULT_OFFICE_CLOSE}
        for i in range(7)
    ]
    
    # Get events grouped by day
    cursor.execute('SELECT * FROM events WHERE schedule_id = ?', (schedule_id,))
    events_by_day = {i: [] for i in range(7)}
    for row in cursor.fetchall():
        if row['event_text']:
            events_by_day[row['day_index']].append(row['event_text'])
    
    # Get employee notes for this week
    cursor.execute(
        'SELECT employee_id, note FROM employee_notes WHERE week_start = ?',
        (week_start_str,)
    )
    notes = {row['employee_id']: row['note'] for row in cursor.fetchall()}
    
    # Build employee data with shifts
    def build_employee(emp):
        return {
            'id': emp['id'],
            'name': emp['name'],
            'phone': emp['phone'] or '',
            'active': bool(emp['active']),
            'shifts': [shifts.get((emp['id'], i)) for i in range(7)],
            'note': notes.get(emp['id'], '')
        }
    
    managers = [build_employee(e) for e in employees if e['section'] == 'manager']
    zak_list = [e for e in employees if e['section'] == 'zak']
    zak = build_employee(zak_list[0]) if zak_list else None
    staff = [build_employee(e) for e in employees if e['section'] == 'staff']

    cursor.execute('''
        SELECT id, name, phone, section FROM employees
        WHERE active = 0
        ORDER BY section, name
    ''')
    hidden_employees = [dict(row) for row in cursor.fetchall()]

    return {
        'weekTitle': schedule['week_title'],
        'weekStart': schedule['week_start'],
        'version': schedule['version'],
        'days': get_week_dates(week_start_str),
        'managers': managers,
        'zakReilly': zak,
        'employees': staff,
        'hiddenEmployees': hidden_employees,
        'officeHours': office_hours,
        'events': [events_by_day[i] for i in range(7)]
    }


# =============================================================================
# ERROR HANDLERS
# =============================================================================

def register_error_handlers(app):
    """Register error handlers"""
    
    @app.errorhandler(400)
    def bad_request(e):
        return jsonify({'error': 'Bad request', 'message': str(e)}), 400
    
    @app.errorhandler(404)
    def not_found(e):
        return jsonify({'error': 'Not found', 'message': str(e)}), 404
    
    @app.errorhandler(500)
    def internal_error(e):
        logging.error(f"Internal error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# =============================================================================
# PDF HELPER
# =============================================================================

def schedule_pdf_filename(data):
    """Build the download filename for a schedule PDF: schedule_<d1>-<yy>_to_<d2>-<yy>.pdf"""
    d1 = data['days'][0]['date'].replace('/', '-')
    d2 = data['days'][6]['date'].replace('/', '-')
    year = data['weekTitle'].split(', ')[-1][-2:]
    return f"schedule_{d1}-{year}_to_{d2}-{year}.pdf"


def build_schedule_pdf_fpdf(data):
    """Render a schedule PDF with FPDF. Returns (pdf_bytes, filename).

    Fallback generator used when headless Chromium is unavailable.
    """
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=10)

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, f"Ice Line Office Schedule - Week of {data['weekTitle']}", ln=True, align='C')
    pdf.ln(3)

    name_col_w = 50
    day_col_w = 26
    sub_col_w = 13
    hours_col_w = 16
    row_h = 7

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(217, 217, 217)
    pdf.cell(name_col_w, row_h, 'Employee', 1, 0, 'C', True)
    for day in data['days']:
        pdf.cell(day_col_w, row_h, f"{day['name']} {day['date']}", 1, 0, 'C', True)
    pdf.cell(hours_col_w, row_h, 'Hours', 1, 1, 'C', True)

    pdf.set_font('Helvetica', 'B', 7)
    pdf.cell(name_col_w, row_h, '', 1, 0, 'C', True)
    for _ in data['days']:
        pdf.cell(sub_col_w, row_h, 'In', 1, 0, 'C', True)
        pdf.cell(sub_col_w, row_h, 'Out', 1, 0, 'C', True)
    pdf.cell(hours_col_w, row_h, '', 1, 1, 'C', True)

    def write_employee_row(emp, fill_rgb=None):
        pdf.set_font('Helvetica', 'B' if fill_rgb else '', 8)
        if fill_rgb:
            pdf.set_fill_color(*fill_rgb)

        name = emp['name']
        if emp.get('phone'):
            name += f"  {emp['phone']}"
        pdf.cell(name_col_w, row_h, name, 1, 0, 'L', bool(fill_rgb))

        pdf.set_font('Helvetica', '', 8)
        for shift in emp['shifts']:
            in_val = shift['in'] if shift and shift.get('in') else ''
            out_val = shift['out'] if shift and shift.get('out') else ''
            pdf.cell(sub_col_w, row_h, str(in_val), 1, 0, 'C', bool(fill_rgb))
            pdf.cell(sub_col_w, row_h, str(out_val), 1, 0, 'C', bool(fill_rgb))

        total = 0
        for shift in emp['shifts']:
            if shift and shift.get('in') and shift.get('out'):
                if shift['in'] == '-' or shift['out'] == '-':
                    continue
                in_mins = parse_time_to_minutes(shift['in'])
                out_mins = parse_time_to_minutes(shift['out'])
                if shift['out'] == 'CLOSE':
                    out_mins = parse_time_to_minutes('10:00 PM')
                if in_mins is not None and out_mins is not None and out_mins > in_mins:
                    total += (out_mins - in_mins) / 60

        hours_str = f"{total:.1f}" if total > 0 else '-'
        pdf.set_font('Helvetica', 'B' if total > 40 else '', 8)
        if total > 40:
            pdf.set_text_color(255, 0, 0)
        pdf.cell(hours_col_w, row_h, hours_str, 1, 1, 'C', bool(fill_rgb))
        pdf.set_text_color(0, 0, 0)

    for emp in data['managers']:
        write_employee_row(emp, (255, 255, 0))
    pdf.ln(2)

    if data['zakReilly']:
        write_employee_row(data['zakReilly'], (146, 208, 80))
    pdf.ln(2)

    for emp in data['employees']:
        write_employee_row(emp)
    pdf.ln(2)

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(255, 255, 0)
    pdf.cell(name_col_w, row_h, 'Front Office Hours*', 1, 0, 'L', True)
    pdf.set_font('Helvetica', '', 8)
    for oh in data['officeHours']:
        in_val = oh.get('in', '') or ''
        out_val = oh.get('out', '') or ''
        if in_val == 'CLOSED':
            pdf.cell(day_col_w, row_h, 'CLOSED', 1, 0, 'C', True)
        else:
            pdf.cell(sub_col_w, row_h, str(in_val), 1, 0, 'C', True)
            pdf.cell(sub_col_w, row_h, str(out_val), 1, 0, 'C', True)
    pdf.cell(hours_col_w, row_h, '', 1, 1, 'C', True)

    pdf.set_font('Helvetica', '', 7)
    pdf.cell(name_col_w, row_h, '* Hours are subject to change', 0, 0, 'L')
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_text_color(255, 0, 0)
    pdf.cell(0, row_h, 'IF UNABLE TO WORK A SCHEDULED SHIFT YOU MUST FIND A REPLACEMENT', 0, 1, 'C')
    pdf.set_text_color(0, 0, 0)

    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(255, 255, 0)
    pdf.cell(name_col_w, row_h, 'Special Events:', 1, 0, 'L', True)
    pdf.set_font('Helvetica', '', 8)
    for events in data['events']:
        event_text = ', '.join(events) if events else ''
        pdf.cell(day_col_w, row_h, event_text, 1, 0, 'C', False)
    pdf.cell(hours_col_w, row_h, '', 1, 1, 'C', False)

    filename = schedule_pdf_filename(data)

    output = BytesIO()
    pdf.output(output)
    output.seek(0)
    return output.read(), filename


def build_schedule_pdf_chromium(week_start, base_url, timeout_ms):
    """Render the live schedule page to PDF with headless Chromium.

    Returns PDF bytes. Raises on any failure (caller falls back to FPDF).
    """
    from playwright.sync_api import sync_playwright

    # A4 landscape printable area in inches, minus 0.25in margins.
    # landscape=True (set on page.pdf below) rotates A4, so the long edge
    # (11.69in) is the width and the short edge (8.27in) is the height.
    margin_in = 0.25
    page_w_in = 11.69 - 2 * margin_in
    page_h_in = 8.27 - 2 * margin_in
    url = f"{base_url}/?print=1&week={week_start}"

    with sync_playwright() as p:
        browser = p.chromium.launch(args=['--no-sandbox'])
        try:
            page = browser.new_page()
            page.emulate_media(media='print')
            # The readiness flag (set by the page after it renders AND its images
            # load) is the authoritative "safe to print" signal, so a plain load
            # wait plus the flag wait is sufficient. timeout_ms applies to each
            # wait independently.
            page.goto(url, timeout=timeout_ms)
            page.wait_for_function('window.__printReady === true', timeout=timeout_ms)

            # Measure full rendered content at scale 1 (CSS px @ 96dpi).
            metrics = page.evaluate(
                "() => ({ w: document.documentElement.scrollWidth,"
                " h: document.documentElement.scrollHeight })"
            )
            if not metrics['w'] or not metrics['h']:
                raise RuntimeError(f"Print page reported zero dimensions: {metrics}")
            content_w_in = metrics['w'] / 96.0
            content_h_in = metrics['h'] / 96.0

            scale = min(page_w_in / content_w_in, page_h_in / content_h_in, 1.0)
            scale = max(scale, 0.1)  # Playwright clamps to [0.1, 2.0].

            pdf_bytes = page.pdf(
                format='A4',
                landscape=True,
                scale=scale,
                print_background=True,
                margin={
                    'top': f'{margin_in}in', 'bottom': f'{margin_in}in',
                    'left': f'{margin_in}in', 'right': f'{margin_in}in',
                },
            )
            return pdf_bytes
        finally:
            browser.close()


# =============================================================================
# ROUTES
# =============================================================================

def register_routes(app):
    """Register all application routes"""
    
    @app.teardown_appcontext
    def teardown(exception):
        close_db()
    
    # CORS removed (April 2026): API_BASE is empty (same-origin); the wildcard
    # was exposing write endpoints to any browser tab. Add a narrow allowlist
    # here if a real cross-origin client ever needs access.


    # -------------------------------------------------------------------------
    # Pages
    # -------------------------------------------------------------------------
    
    def _static_version(*relpaths):
        """Combined mtime of static assets, used to bust browser/CDN caches on deploy."""
        latest = 0
        for rel in relpaths:
            path = os.path.join(app.static_folder, rel)
            try:
                latest = max(latest, int(os.path.getmtime(path)))
            except OSError:
                pass
        return str(latest)

    @app.route('/')
    @manager_required
    def index():
        return render_template(
            'index.html',
            static_v=_static_version('js/main.js', 'css/main.css'),
            backup_import_enabled=Config.BACKUP_IMPORT_ENABLED,
        )

    @app.route('/employee')
    @app.route('/employee/')
    @employee_required
    def employee_portal():
        """Employee portal - read-only schedule view"""
        return render_template(
            'employee_portal.html',
            static_v=_static_version('js/employee_portal.js', 'css/employee_portal.css'),
        )

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            username = (request.form.get('username') or '').strip()
            password = request.form.get('password') or ''
            emp_id, reason = verify_login(get_db(), username, password, 'manager')
            if emp_id:
                login_user(emp_id, 'manager', permanent=False)
                return redirect('/')
            msg = ('Account temporarily locked. Try again later.'
                   if reason == 'locked' else 'Invalid credentials.')
            return render_template('login.html', error=msg), 401
        return render_template('login.html', error=None)

    @app.route('/employee/login', methods=['GET', 'POST'])
    def employee_login():
        if request.method == 'POST':
            identifier = (request.form.get('name') or '').strip()
            pin = request.form.get('pin') or ''
            emp_id, reason = verify_login(get_db(), identifier, pin, 'employee')
            if emp_id:
                login_user(emp_id, 'employee', permanent=True)
                return redirect('/employee')
            msg = ('Account temporarily locked. Try again later.'
                   if reason == 'locked' else 'Invalid name or PIN.')
            return render_template('employee_login.html', error=msg), 401
        return render_template('employee_login.html', error=None)

    @app.route('/logout', methods=['GET', 'POST'])
    def logout():
        was_manager = session.get('role') == 'manager'
        logout_user()
        return redirect('/login' if was_manager else '/employee/login')

    # -------------------------------------------------------------------------
    # Schedule API
    # -------------------------------------------------------------------------
    
    @app.route('/api/current-week', methods=['GET'])
    @employee_required
    def get_current_week():
        """Get the current schedule week start date"""
        today = date.today()
        return jsonify({
            'weekStart': get_current_week_start(),
            'today': today.isoformat(),
            'todayName': today.strftime('%A')
        })
    
    @app.route('/api/schedule/<week_start>', methods=['GET'])
    @employee_required
    def get_schedule(week_start):
        """Get schedule for a specific week"""
        try:
            return jsonify(build_schedule_response(week_start))
        except Exception as e:
            logging.error(f"Error getting schedule: {e}")
            return jsonify({'error': str(e)}), 400
    
    @app.route('/api/schedule/<week_start>', methods=['POST'])
    @manager_required
    def save_schedule(week_start):
        """Save/update schedule for a specific week with optimistic concurrency.

        Client must send `base_version` matching the version it last received.
        Returns 409 if another save bumped the version since then; client
        should reload and surface the conflict to the user.
        """
        try:
            data = request.json
            base_version = data.get('base_version')

            schedule = get_or_create_schedule(week_start)
            schedule_id = schedule['id']

            db = get_db()
            cursor = db.cursor()

            # Optimistic concurrency: atomic version bump. If base_version
            # doesn't match current, this UPDATE affects 0 rows -> conflict.
            # base_version may be None for legacy clients during the rollout
            # window; in that case we accept the save but log it.
            if base_version is not None:
                cursor.execute(
                    '''UPDATE schedules
                       SET version = version + 1, updated_at = CURRENT_TIMESTAMP
                       WHERE id = ? AND version = ?''',
                    (schedule_id, base_version)
                )
                if cursor.rowcount == 0:
                    cursor.execute('SELECT version FROM schedules WHERE id = ?', (schedule_id,))
                    current = cursor.fetchone()
                    db.rollback()
                    return jsonify({
                        'error': 'conflict',
                        'message': 'Schedule was updated elsewhere. Reload to see the latest.',
                        'current_version': current['version'] if current else None,
                    }), 409
            else:
                cursor.execute(
                    '''UPDATE schedules
                       SET version = version + 1, updated_at = CURRENT_TIMESTAMP
                       WHERE id = ?''',
                    (schedule_id,)
                )
                logging.warning(f"save_schedule: legacy client (no base_version) for week {week_start}")

            # Update shifts
            if 'shifts' in data:
                cursor.execute('DELETE FROM shifts WHERE schedule_id = ?', (schedule_id,))
                for shift in data['shifts']:
                    if shift.get('in') or shift.get('out'):
                        cursor.execute(
                            '''INSERT INTO shifts
                               (schedule_id, employee_id, day_index, time_in, time_out)
                               VALUES (?, ?, ?, ?, ?)''',
                            (schedule_id, shift['employee_id'], shift['day_index'],
                             shift.get('in'), shift.get('out'))
                        )

            # Update office hours
            if 'officeHours' in data:
                cursor.execute('DELETE FROM office_hours WHERE schedule_id = ?', (schedule_id,))
                for i, oh in enumerate(data['officeHours']):
                    cursor.execute(
                        '''INSERT INTO office_hours
                           (schedule_id, day_index, time_in, time_out)
                           VALUES (?, ?, ?, ?)''',
                        (schedule_id, i, oh.get('in'), oh.get('out'))
                    )

            # Update events
            if 'events' in data:
                cursor.execute('DELETE FROM events WHERE schedule_id = ?', (schedule_id,))
                for i, day_events in enumerate(data['events']):
                    if day_events:
                        for event_text in day_events:
                            if event_text:
                                cursor.execute(
                                    '''INSERT INTO events
                                       (schedule_id, day_index, event_text)
                                       VALUES (?, ?, ?)''',
                                    (schedule_id, i, event_text)
                                )

            db.commit()

            cursor.execute('SELECT version FROM schedules WHERE id = ?', (schedule_id,))
            new_version = cursor.fetchone()['version']
            return jsonify({'success': True, 'message': 'Schedule saved', 'version': new_version})
        
        except Exception as e:
            import traceback
            logging.error(f"Error saving schedule: {e}")
            logging.error(traceback.format_exc())
            return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 400
    
    @app.route('/api/schedule/<week_start>/shift', methods=['POST'])
    @manager_required
    def update_shift(week_start):
        """Update a single shift"""
        try:
            data = request.json
            schedule = get_or_create_schedule(week_start)
            
            db = get_db()
            cursor = db.cursor()
            
            # Delete existing shift
            cursor.execute(
                '''DELETE FROM shifts 
                   WHERE schedule_id = ? AND employee_id = ? AND day_index = ?''',
                (schedule['id'], data['employee_id'], data['day_index'])
            )
            
            # Insert new shift if values provided
            if data.get('in') or data.get('out'):
                cursor.execute(
                    '''INSERT INTO shifts 
                       (schedule_id, employee_id, day_index, time_in, time_out) 
                       VALUES (?, ?, ?, ?, ?)''',
                    (schedule['id'], data['employee_id'], data['day_index'],
                     data.get('in'), data.get('out'))
                )
            
            db.commit()
            return jsonify({'success': True})
        
        except Exception as e:
            logging.error(f"Error updating shift: {e}")
            return jsonify({'error': str(e)}), 400
    
    # -------------------------------------------------------------------------
    # Employee API
    # -------------------------------------------------------------------------
    
    @app.route('/api/employees', methods=['GET'])
    @manager_required
    def get_employees():
        """Get all active employees"""
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            '''SELECT * FROM employees 
               WHERE active = 1 
               ORDER BY section, sort_order'''
        )
        return jsonify([dict(row) for row in cursor.fetchall()])
    
    @app.route('/api/employees', methods=['POST'])
    @manager_required
    def add_employee():
        """Add a new employee"""
        try:
            data = request.json
            
            if not data.get('name'):
                return jsonify({'error': 'Name is required'}), 400
            
            section = data.get('section', 'staff')
            if section not in ('manager', 'zak', 'staff'):
                return jsonify({'error': 'Invalid section'}), 400
            
            db = get_db()
            cursor = db.cursor()
            
            # Get next sort order
            cursor.execute(
                'SELECT COALESCE(MAX(sort_order), 0) + 1 FROM employees WHERE section = ?',
                (section,)
            )
            next_order = cursor.fetchone()[0]
            
            cursor.execute(
                '''INSERT INTO employees (name, phone, section, sort_order) 
                   VALUES (?, ?, ?, ?)''',
                (data['name'], data.get('phone', ''), section, next_order)
            )
            
            emp_id = cursor.lastrowid
            cursor.execute('SELECT * FROM employees WHERE id = ?', (emp_id,))
            employee = dict(cursor.fetchone())
            
            db.commit()
            return jsonify(employee), 201
        
        except Exception as e:
            logging.error(f"Error adding employee: {e}")
            return jsonify({'error': str(e)}), 400
    
    @app.route('/api/employees/<int:emp_id>', methods=['PUT'])
    @manager_required
    def update_employee(emp_id):
        """Update an employee"""
        try:
            data = request.json
            db = get_db()
            cursor = db.cursor()
            
            # Build update query dynamically
            updates = []
            values = []
            
            for field in ['name', 'phone', 'sort_order']:
                if field in data:
                    updates.append(f'{field} = ?')
                    values.append(data[field])
            
            if updates:
                updates.append('updated_at = CURRENT_TIMESTAMP')
                values.append(emp_id)
                cursor.execute(
                    f'UPDATE employees SET {", ".join(updates)} WHERE id = ?',
                    values
                )
            
            cursor.execute('SELECT * FROM employees WHERE id = ?', (emp_id,))
            row = cursor.fetchone()
            
            if not row:
                return jsonify({'error': 'Employee not found'}), 404
            
            db.commit()
            return jsonify(dict(row))
        
        except Exception as e:
            logging.error(f"Error updating employee: {e}")
            return jsonify({'error': str(e)}), 400
    
    @app.route('/api/employees/<int:emp_id>', methods=['DELETE'])
    @manager_required
    def delete_employee(emp_id):
        """Hide an employee from new schedules. Past weeks where they had shifts
        still show them; the Hidden Employees panel can bring them back."""
        try:
            db = get_db()
            cursor = db.cursor()
            cursor.execute(
                'UPDATE employees SET active = 0, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                (emp_id,)
            )
            db.commit()
            return jsonify({'success': True})

        except Exception as e:
            logging.error(f"Error hiding employee: {e}")
            return jsonify({'error': str(e)}), 400

    @app.route('/api/employees/<int:emp_id>/restore', methods=['POST'])
    @manager_required
    def restore_employee(emp_id):
        """Bring a hidden employee back into active scheduling."""
        try:
            db = get_db()
            cursor = db.cursor()
            cursor.execute(
                'UPDATE employees SET active = 1, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                (emp_id,)
            )
            cursor.execute('SELECT * FROM employees WHERE id = ?', (emp_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({'error': 'Employee not found'}), 404
            db.commit()
            return jsonify(dict(row))

        except Exception as e:
            logging.error(f"Error restoring employee: {e}")
            return jsonify({'error': str(e)}), 400
    
    # -------------------------------------------------------------------------
    # Employee Notes API
    # -------------------------------------------------------------------------
    
    @app.route('/api/notes/<week_start>/<int:emp_id>', methods=['GET'])
    @manager_required
    def get_note(week_start, emp_id):
        """Get note for employee for a specific week"""
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            'SELECT note FROM employee_notes WHERE week_start = ? AND employee_id = ?',
            (week_start, emp_id)
        )
        row = cursor.fetchone()
        return jsonify({'note': row['note'] if row else ''})
    
    @app.route('/api/notes/<week_start>/<int:emp_id>', methods=['POST'])
    @manager_required
    def save_note(week_start, emp_id):
        """Save note for employee for a specific week"""
        try:
            data = request.json
            note = data.get('note', '').strip()
            
            db = get_db()
            cursor = db.cursor()
            
            if note:
                cursor.execute(
                    '''INSERT INTO employee_notes (employee_id, week_start, note) 
                       VALUES (?, ?, ?)
                       ON CONFLICT(employee_id, week_start) 
                       DO UPDATE SET note = ?, updated_at = CURRENT_TIMESTAMP''',
                    (emp_id, week_start, note, note)
                )
            else:
                cursor.execute(
                    'DELETE FROM employee_notes WHERE employee_id = ? AND week_start = ?',
                    (emp_id, week_start)
                )
            
            db.commit()
            return jsonify({'success': True})
        
        except Exception as e:
            logging.error(f"Error saving note: {e}")
            return jsonify({'error': str(e)}), 400
    
    # -------------------------------------------------------------------------
    # Export API
    # -------------------------------------------------------------------------
    
    @app.route('/api/schedule/<week_start>/export', methods=['GET'])
    @manager_required
    def export_schedule(week_start):
        """Export schedule to Excel with formatting"""
        try:
            data = build_schedule_response(week_start)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Schedule"
            
            # Styles
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            bold_font = Font(bold=True)
            red_bold_font = Font(bold=True, color="FF0000")
            title_font = Font(bold=True, size=14)
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            # Title row
            ws.merge_cells('A1:O1')
            ws['A1'] = f"Ice Line Office Schedule for week of {data['weekTitle']}"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align
            
            # Day headers
            col = 2
            for day in data['days']:
                for r in [2, 3]:
                    for c in [col, col + 1]:
                        ws.cell(row=r, column=c).fill = gray_fill
                        ws.cell(row=r, column=c).alignment = center_align
                        ws.cell(row=r, column=c).border = thin_border
                ws.cell(row=2, column=col, value=day['name'])
                ws.cell(row=2, column=col + 1, value=day['date'])
                ws.cell(row=3, column=col, value='In')
                ws.cell(row=3, column=col + 1, value='Out')
                col += 2
            
            current_row = 4
            
            def write_employee_row(emp, row, fill=None):
                name = emp['name'] + ('     ' + emp['phone'] if emp.get('phone') else '')
                ws.cell(row=row, column=1, value=name).alignment = left_align
                ws.cell(row=row, column=1).border = thin_border
                if fill:
                    ws.cell(row=row, column=1).fill = fill
                    ws.cell(row=row, column=1).font = bold_font
                
                col = 2
                for shift in emp['shifts']:
                    in_val = shift['in'] if shift else ''
                    out_val = shift['out'] if shift else ''
                    for c, v in [(col, in_val), (col + 1, out_val)]:
                        ws.cell(row=row, column=c, value=v).alignment = center_align
                        ws.cell(row=row, column=c).border = thin_border
                        if fill:
                            ws.cell(row=row, column=c).fill = fill
                    col += 2
            
            # Managers
            for emp in data['managers']:
                write_employee_row(emp, current_row, yellow_fill)
                current_row += 1
            
            current_row += 1  # Empty row
            
            # Zak
            if data['zakReilly']:
                write_employee_row(data['zakReilly'], current_row, green_fill)
                current_row += 1
            
            current_row += 4  # Empty rows
            
            # Staff
            for emp in data['employees']:
                write_employee_row(emp, current_row, None)
                current_row += 1
            
            # Office Hours
            ws.cell(row=current_row, column=1, value='Front Office Hours*')
            ws.cell(row=current_row, column=1).fill = yellow_fill
            ws.cell(row=current_row, column=1).font = bold_font
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=1).border = thin_border
            
            col = 2
            for oh in data['officeHours']:
                for c, v in [(col, oh['in']), (col + 1, oh['out'])]:
                    ws.cell(row=current_row, column=c, value=v).fill = yellow_fill
                    ws.cell(row=current_row, column=c).alignment = center_align
                    ws.cell(row=current_row, column=c).border = thin_border
                col += 2
            current_row += 1
            
            # Notice row
            ws.cell(row=current_row, column=1, value='* Hours are subject to change')
            ws.cell(row=current_row, column=3, 
                    value='IF UNABLE TO WORK A SCHEDULED SHIFT YOU MUST FIND A REPLACEMENT')
            ws.cell(row=current_row, column=3).fill = green_fill
            ws.cell(row=current_row, column=3).font = red_bold_font
            current_row += 1
            
            # Events row
            ws.cell(row=current_row, column=1, value='Special Events:')
            ws.cell(row=current_row, column=1).fill = yellow_fill
            ws.cell(row=current_row, column=1).font = bold_font
            col = 2
            for events in data['events']:
                ws.cell(row=current_row, column=col, 
                        value=', '.join(events) if events else '').alignment = center_align
                col += 2
            
            # Column widths
            ws.column_dimensions['A'].width = 35
            for c in range(2, 16):
                ws.column_dimensions[get_column_letter(c)].width = 10
            
            # Save to buffer
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Generate filename
            d1 = data['days'][0]['date'].replace('/', '-')
            d2 = data['days'][6]['date'].replace('/', '-')
            year = data['weekTitle'].split(', ')[-1][-2:]
            filename = f"schedule_{d1}-{year}_to_{d2}-{year}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        except Exception as e:
            logging.error(f"Error exporting schedule: {e}")
            return jsonify({'error': str(e)}), 400

    # -------------------------------------------------------------------------
    # PDF Export API
    # -------------------------------------------------------------------------

    @app.route('/api/schedule/<week_start>/export-pdf', methods=['GET'])
    @manager_required
    def export_schedule_pdf(week_start):
        """Export schedule to PDF. Prefers headless Chromium (pixel-matches the
        on-screen print view); falls back to the FPDF generator if Chromium is
        unavailable or errors."""
        try:
            # Validate the date format up front: week_start is interpolated into
            # the URL handed to headless Chromium.
            try:
                datetime.strptime(week_start, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid week_start format (expected YYYY-MM-DD)'}), 400

            data = build_schedule_response(week_start)
            filename = schedule_pdf_filename(data)

            pdf_bytes = None
            try:
                pdf_bytes = build_schedule_pdf_chromium(
                    week_start,
                    Config.PDF_RENDER_BASE_URL,
                    Config.PDF_RENDER_TIMEOUT_MS,
                )
            except Exception as e:
                logging.warning(f"Chromium PDF render failed, using FPDF fallback: {e}")
                pdf_bytes, filename = build_schedule_pdf_fpdf(data)

            return send_file(
                BytesIO(pdf_bytes),
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename,
            )
        except Exception as e:
            logging.error(f"Error exporting PDF: {e}")
            return jsonify({'error': str(e)}), 400

    @app.route('/api/timeoff-calendars', methods=['GET'])
    @manager_required
    def export_timeoff_calendars():
        """Generate blank printable monthly calendars for staff time-off requests.

        One page per month. Each day cell has AM / PM / All Day checkboxes plus a
        write-in line for specific times. Defaults to the next 3 months; override
        with ?start=YYYY-MM and ?months=N.
        """
        try:
            # Determine the starting month (default: next calendar month).
            start_param = request.args.get('start')
            if start_param:
                start_year, start_month = (int(x) for x in start_param.split('-'))
            else:
                today = date.today()
                start_year = today.year + (1 if today.month == 12 else 0)
                start_month = 1 if today.month == 12 else today.month + 1

            months = max(1, min(12, request.args.get('months', default=3, type=int)))

            # Build the list of (year, month) pairs.
            month_list = []
            y, m = start_year, start_month
            for _ in range(months):
                month_list.append((y, m))
                m += 1
                if m > 12:
                    m, y = 1, y + 1

            pdf = FPDF(orientation='P', unit='mm', format='Letter')
            pdf.set_auto_page_break(auto=False)

            logo_path = os.path.join(app.static_folder, 'Ice_Line_Logo.png')
            has_logo = os.path.exists(logo_path)

            page_w = 215.9      # Letter width (mm)
            margin = 10
            usable_w = page_w - 2 * margin
            col_w = usable_w / 7
            weekday_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

            def draw_checkbox(x, y, label, size=3.2):
                pdf.rect(x, y, size, size)
                pdf.set_xy(x + size + 1, y - 0.6)
                pdf.set_font('Helvetica', '', 6.5)
                pdf.cell(col_w / 2, size + 1.2, label, 0, 0, 'L')

            for (year, month) in month_list:
                pdf.add_page()

                # ---- Header ----
                if has_logo:
                    pdf.image(logo_path, x=margin, y=8, w=22)
                pdf.set_xy(margin, 9)
                pdf.set_font('Helvetica', 'B', 16)
                pdf.cell(0, 8, 'Ice Line Quad Rinks - Time-Off Request', 0, 1, 'C')
                pdf.set_x(margin)
                pdf.set_font('Helvetica', 'B', 13)
                pdf.cell(0, 7, f"{cal.month_name[month]} {year}", 0, 1, 'C')

                # Name / submitted lines
                pdf.ln(1)
                pdf.set_x(margin)
                pdf.set_font('Helvetica', '', 10)
                pdf.cell(usable_w * 0.6, 7, 'Name: ______________________________', 0, 0, 'L')
                pdf.cell(usable_w * 0.4, 7, 'Date submitted: __________________', 0, 1, 'L')

                # Instructions box
                pdf.ln(1)
                instr_x, instr_y = margin, pdf.get_y()
                instr_h = 25
                pdf.set_fill_color(238, 242, 248)
                pdf.set_draw_color(150, 150, 150)
                pdf.rect(instr_x, instr_y, usable_w, instr_h, 'DF')
                pdf.set_xy(instr_x + 2, instr_y + 1.5)
                pdf.set_font('Helvetica', 'B', 9)
                pdf.cell(0, 4.5, 'On each day you need off, mark one:', 0, 1, 'L')
                pdf.set_font('Helvetica', '', 8.5)
                bullets = [
                    ('All Day', 'check this box for the whole day off'),
                    ('AM  or  PM', 'check one for just the morning or just the afternoon/evening'),
                    ('Line at the bottom', 'write exact hours if needed (e.g. "off after 2:00")'),
                ]
                for label, desc in bullets:
                    pdf.set_x(instr_x + 4)
                    pdf.set_font('Helvetica', 'B', 8.5)
                    pdf.cell(34, 4, f"-  {label}", 0, 0, 'L')
                    pdf.set_font('Helvetica', '', 8.5)
                    pdf.cell(0, 4, desc, 0, 1, 'L')
                pdf.set_x(instr_x + 2)
                pdf.set_font('Helvetica', 'I', 8)
                pdf.cell(0, 4, 'Return this sheet to the office by the posted deadline.', 0, 1, 'L')
                pdf.set_draw_color(0, 0, 0)

                # ---- Calendar grid ----
                grid_top = instr_y + instr_h + 3
                bottom_margin = 10
                wd_header_h = 6

                weeks = cal.Calendar(firstweekday=6).monthdayscalendar(year, month)
                page_h = 279.4  # Letter height (mm)
                grid_h = page_h - grid_top - bottom_margin
                row_h = (grid_h - wd_header_h) / len(weeks)

                # Weekday header
                pdf.set_xy(margin, grid_top)
                pdf.set_font('Helvetica', 'B', 9)
                pdf.set_fill_color(30, 58, 95)
                pdf.set_text_color(255, 255, 255)
                for name in weekday_names:
                    pdf.cell(col_w, wd_header_h, name, 1, 0, 'C', True)
                pdf.set_text_color(0, 0, 0)

                # Day cells
                for w_idx, week in enumerate(weeks):
                    cell_y = grid_top + wd_header_h + w_idx * row_h
                    for d_idx, daynum in enumerate(week):
                        cell_x = margin + d_idx * col_w
                        if daynum == 0:
                            pdf.set_fill_color(245, 245, 245)
                            pdf.rect(cell_x, cell_y, col_w, row_h, 'DF')
                            pdf.set_fill_color(255, 255, 255)
                            continue
                        pdf.rect(cell_x, cell_y, col_w, row_h)
                        # Day number
                        pdf.set_xy(cell_x + 1, cell_y + 0.8)
                        pdf.set_font('Helvetica', 'B', 11)
                        pdf.cell(col_w - 2, 5, str(daynum), 0, 0, 'L')
                        # Checkboxes
                        cb_y = cell_y + 8
                        draw_checkbox(cell_x + 2, cb_y, 'AM')
                        draw_checkbox(cell_x + col_w / 2 + 1, cb_y, 'PM')
                        draw_checkbox(cell_x + 2, cb_y + 6, 'All Day')
                        # Write-in line for specific times
                        line_y = cell_y + row_h - 3.5
                        pdf.set_draw_color(180, 180, 180)
                        pdf.line(cell_x + 2, line_y, cell_x + col_w - 2, line_y)
                        pdf.set_draw_color(0, 0, 0)

            filename = (
                f"timeoff_{cal.month_abbr[month_list[0][1]]}-"
                f"{cal.month_abbr[month_list[-1][1]]}_{month_list[0][0]}.pdf"
            )

            output = BytesIO()
            pdf.output(output)
            output.seek(0)
            return send_file(output, mimetype='application/pdf',
                             as_attachment=True, download_name=filename)

        except Exception as e:
            logging.error(f"Error generating time-off calendars: {e}")
            return jsonify({'error': str(e)}), 400

    # -------------------------------------------------------------------------
    # Database Export/Import API
    # -------------------------------------------------------------------------

    @app.route('/api/backup/export', methods=['GET'])
    @manager_required
    def export_database():
        """Export entire database to JSON for backup"""
        try:
            db = get_db()
            cursor = db.cursor()
            
            backup_data = {
                'export_date': datetime.now().isoformat(),
                'version': '1.0',
                'employees': [],
                'schedules': [],
                'shifts': [],
                'office_hours': [],
                'events': [],
                'employee_notes': []
            }
            
            # Export employees
            cursor.execute('SELECT * FROM employees')
            for row in cursor.fetchall():
                backup_data['employees'].append(dict(row))
            
            # Export schedules
            cursor.execute('SELECT * FROM schedules')
            for row in cursor.fetchall():
                backup_data['schedules'].append(dict(row))
            
            # Export shifts
            cursor.execute('SELECT * FROM shifts')
            for row in cursor.fetchall():
                backup_data['shifts'].append(dict(row))
            
            # Export office hours
            cursor.execute('SELECT * FROM office_hours')
            for row in cursor.fetchall():
                backup_data['office_hours'].append(dict(row))
            
            # Export events
            cursor.execute('SELECT * FROM events')
            for row in cursor.fetchall():
                backup_data['events'].append(dict(row))
            
            # Export employee notes
            cursor.execute('SELECT * FROM employee_notes')
            for row in cursor.fetchall():
                backup_data['employee_notes'].append(dict(row))
            
            # Create JSON file response
            import json
            output = BytesIO()
            output.write(json.dumps(backup_data, indent=2).encode('utf-8'))
            output.seek(0)
            
            filename = f"schedule_backup_{datetime.now().strftime('%Y-%m-%d_%H%M')}.json"
            
            return send_file(
                output,
                mimetype='application/json',
                as_attachment=True,
                download_name=filename
            )
        
        except Exception as e:
            logging.error(f"Error exporting database: {e}")
            return jsonify({'error': str(e)}), 400
    
    @app.route('/api/backup/import', methods=['POST'])
    @manager_required
    def import_database():
        """Import database from JSON backup. Disabled by default — this
        endpoint replaces the entire DB and there is no auth on the app yet.
        Set BACKUP_IMPORT_ENABLED=true in the LXC env to re-enable temporarily
        when running a restore, then unset it."""
        if not Config.BACKUP_IMPORT_ENABLED:
            return jsonify({
                'error': 'Backup import is disabled',
                'message': 'Set BACKUP_IMPORT_ENABLED=true and restart the service to re-enable temporarily.'
            }), 403
        try:
            if 'file' not in request.files:
                return jsonify({'error': 'No file provided'}), 400
            
            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            import json
            backup_data = json.load(file)
            
            # Validate backup structure
            required_keys = ['employees', 'schedules', 'shifts', 'office_hours', 'events']
            for key in required_keys:
                if key not in backup_data:
                    return jsonify({'error': f'Invalid backup file: missing {key}'}), 400
            
            db = get_db()
            cursor = db.cursor()
            
            # Clear existing data (in reverse order of dependencies)
            cursor.execute('DELETE FROM employee_notes')
            cursor.execute('DELETE FROM events')
            cursor.execute('DELETE FROM office_hours')
            cursor.execute('DELETE FROM shifts')
            cursor.execute('DELETE FROM schedules')
            cursor.execute('DELETE FROM employees')
            
            # Import employees
            for emp in backup_data['employees']:
                cursor.execute('''
                    INSERT INTO employees (id, name, phone, section, sort_order, active, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (emp['id'], emp['name'], emp.get('phone', ''), emp['section'], 
                      emp.get('sort_order', 1), emp.get('active', 1),
                      emp.get('created_at'), emp.get('updated_at')))
            
            # Import schedules
            for sched in backup_data['schedules']:
                cursor.execute('''
                    INSERT INTO schedules (id, week_start, week_title, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                ''', (sched['id'], sched['week_start'], sched['week_title'],
                      sched.get('created_at'), sched.get('updated_at')))
            
            # Import shifts
            for shift in backup_data['shifts']:
                cursor.execute('''
                    INSERT INTO shifts (id, schedule_id, employee_id, day_index, time_in, time_out)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (shift['id'], shift['schedule_id'], shift['employee_id'],
                      shift['day_index'], shift.get('time_in'), shift.get('time_out')))
            
            # Import office hours
            for oh in backup_data['office_hours']:
                cursor.execute('''
                    INSERT INTO office_hours (id, schedule_id, day_index, time_in, time_out)
                    VALUES (?, ?, ?, ?, ?)
                ''', (oh['id'], oh['schedule_id'], oh['day_index'],
                      oh.get('time_in'), oh.get('time_out')))
            
            # Import events
            for event in backup_data['events']:
                cursor.execute('''
                    INSERT INTO events (id, schedule_id, day_index, event_text)
                    VALUES (?, ?, ?, ?)
                ''', (event['id'], event['schedule_id'], event['day_index'], event.get('event_text')))
            
            # Import employee notes (if present)
            if 'employee_notes' in backup_data:
                for note in backup_data['employee_notes']:
                    cursor.execute('''
                        INSERT INTO employee_notes (id, employee_id, week_start, note, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (note['id'], note['employee_id'], note['week_start'],
                          note.get('note'), note.get('created_at'), note.get('updated_at')))
            
            db.commit()
            
            return jsonify({
                'success': True,
                'message': 'Database restored successfully',
                'stats': {
                    'employees': len(backup_data['employees']),
                    'schedules': len(backup_data['schedules']),
                    'shifts': len(backup_data['shifts']),
                    'office_hours': len(backup_data['office_hours']),
                    'events': len(backup_data['events']),
                    'employee_notes': len(backup_data.get('employee_notes', []))
                }
            })
        
        except json.JSONDecodeError:
            return jsonify({'error': 'Invalid JSON file'}), 400
        except Exception as e:
            logging.error(f"Error importing database: {e}")
            return jsonify({'error': str(e)}), 400


# =============================================================================
# MAIN
# =============================================================================

app = create_app()

if __name__ == '__main__':
    # threaded=True lets the dev server handle the headless-Chromium self-request
    # (issued while an export-pdf request is in flight) without deadlocking.
    # Production gunicorn (--workers 9) already handles concurrent requests.
    app.run(host='0.0.0.0', port=5001, debug=Config.DEBUG, threaded=True)
